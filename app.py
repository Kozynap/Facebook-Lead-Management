import io
import os
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import db
import importer

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

app = Flask(__name__)
db.init_db()


def _display(iso_str):
    if not iso_str:
        return None
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d %b %Y, %I:%M %p")
    except ValueError:
        return iso_str


def _asset_version(*rel_parts):
    """File mtime used as a cache-busting query param, so a browser can
    never serve a stale static/js or static/css file against newer HTML."""
    path = os.path.join(BASE_DIR, "static", *rel_parts)
    try:
        return str(int(os.path.getmtime(path)))
    except OSError:
        return "0"


@app.route("/")
def index():
    return render_template(
        "index.html",
        status_options=db.STATUS_OPTIONS,
        css_version=_asset_version("css", "style.css"),
        js_version=_asset_version("js", "app.js"),
    )


@app.route("/api/meta")
def api_meta():
    return jsonify({
        "last_created": _display(db.get_last_created()),
        "last_updated": _display(db.get_meta("last_upload_at")),
        "months": db.get_distinct_months(),
        "areas": db.get_area_counts(),
        "unassigned_campaign_count": db.get_unassigned_campaign_count(),
        "status_options": db.STATUS_OPTIONS,
    })


@app.route("/api/leads")
def api_leads():
    return jsonify(db.get_all_leads())


@app.route("/api/campaign-map")
def api_campaign_map_get():
    return jsonify(db.get_campaigns_with_area())


@app.route("/api/campaign-map", methods=["POST"])
def api_campaign_map_set():
    body = request.get_json(force=True, silent=True) or {}
    campaign = body.get("campaign")
    area = body.get("area") or None

    if not campaign:
        return jsonify({"error": "campaign is required"}), 400
    if area is not None and area not in db.AREAS:
        return jsonify({"error": "Invalid area."}), 400

    db.set_campaign_area(campaign, area)
    return jsonify({"campaign": campaign, "area": area})


@app.route("/api/export/tracking")
def export_tracking():
    area = request.args.get("area", "")
    month = request.args.get("month", "all")
    campaign = request.args.get("campaign", "")
    status = request.args.get("status", "all")

    leads = db.get_all_leads()
    filtered = [
        l for l in leads
        if (not area or l["area"] == area)
        and (month == "all" or l["created_month"] == month)
        and (not campaign or l["campaign_name"] == campaign)
        and (status == "all" or l["status"] == status)
    ]

    headers = ["Lead ID", "Full Name", "Email", "Phone", "Street Address", "Status", "Attempts Made", "Remarks"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Tracking"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    def row_values(l):
        return [
            l["id"], l["full_name"], l["email"], l["phone"], l["street_address"],
            l["status"], f"Attempt {l['attempts']}", l["remarks"],
        ]

    for l in filtered:
        ws.append(row_values(l))

    widths = [len(h) for h in headers]
    for l in filtered:
        for i, val in enumerate(row_values(l)):
            widths[i] = max(widths[i], len(str(val or "")))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    name_bits = [
        area or "all-areas",
        campaign or "all-campaigns",
        month if month != "all" else "all-months",
        status if status != "all" else "all-statuses",
    ]
    filename = "client-tracking_" + "_".join(b.strip().replace(" ", "-") for b in name_bits) + ".xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/upload", methods=["POST"])
def api_upload():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files were uploaded."}), 400

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    imported_at = datetime.now().astimezone().isoformat()

    total = {"inserted": 0, "updated": 0, "skipped_rows": 0, "files_processed": 0, "errors": []}
    processed_any = False

    for f in files:
        if not f.filename:
            continue
        if not f.filename.lower().endswith(".zip"):
            total["errors"].append(f"{f.filename}: not a .zip file, skipped")
            continue

        data = f.read()

        safe_name = f"{imported_at.replace(':', '-')}__{os.path.basename(f.filename)}"
        with open(os.path.join(UPLOAD_DIR, safe_name), "wb") as out:
            out.write(data)

        stats = importer.process_zip(data, f.filename, imported_at)
        for key in ("inserted", "updated", "skipped_rows", "files_processed"):
            total[key] += stats[key]
        total["errors"].extend(stats["errors"])
        processed_any = True

    if processed_any:
        db.set_meta("last_upload_at", imported_at)

    return jsonify(total)


@app.route("/api/leads/<path:lead_id>", methods=["PATCH"])
def api_update_lead(lead_id):
    body = request.get_json(force=True, silent=True) or {}
    status = body.get("status")
    remarks = body.get("remarks")

    if status is not None and status not in db.STATUS_OPTIONS:
        return jsonify({"error": "Invalid status value."}), 400

    updated = db.update_lead(lead_id, status=status, remarks=remarks)
    if updated is None:
        return jsonify({"error": "Lead not found."}), 404
    return jsonify(updated)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
